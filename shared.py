# shared.py

# --- env + google clients (matches your .env names) ---
import os, asyncio
import aiosqlite
from pathlib import Path
import pytz
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from dateutil import parser as du_parser
from typing import Dict, List, Tuple, Optional
import time, random, httplib2
from google_auth_httplib2 import AuthorizedHttp
from googleapiclient.errors import HttpError
from pathlib import Path
from typing import Iterable, Mapping, Any
import pandas as pd

ROOT = Path(__file__).resolve().parent
load_dotenv(ROOT / ".env")

TZ_NAME = os.getenv("TZ", "America/Chicago")
TZ = pytz.timezone(TZ_NAME)

CAL_ID = os.getenv("CALENDAR_ID")
KEY_PATH = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "creds/service-account.json")
KEY_PATH = str((ROOT / KEY_PATH).resolve())

EVENT_CHANNEL_ID = int(os.getenv("EVENT_CHANNEL_ID", "0"))
CREATE_FROM_CHANNEL_ID = int(os.getenv("CREATE_FROM_CHANNEL_ID", "0"))
GUILD_ID = int(os.getenv("GUILD_ID", "0"))
DB_PATH = os.getenv("DB_PATH", "data/bot.db")

ROLES_SHEET = os.getenv("ROLES_SHEET_ID")
ROLES_TAB   = os.getenv("ROLES_TAB_NAME", "Roles")
MEMBERS_TAB = os.getenv("MEMBER_TAB_NAME", "Members")
RULES_SHEET = os.getenv("RULES_SHEET_ID")
EVENT_SHEET = os.getenv("EVENT_SHEET_ID")
EVENT_TAB   = os.getenv("EVENT_TAB_NAME", "Events")
YARDSALE_CHANNEL_ID = os.getenv("YARDSALE_CHANNEL_ID", "0")
FOODEATS_CHANNEL_ID = os.getenv("FOODEATS_CHANNEL_ID", "0")

if not CAL_ID:
    raise RuntimeError("CALENDAR_ID is required")
if not os.path.isfile(KEY_PATH):
    raise RuntimeError(f"Service account key not found: {KEY_PATH}")

SCOPES = [
    "https://www.googleapis.com/auth/calendar",
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly"
]
creds = Credentials.from_service_account_file(KEY_PATH, scopes=SCOPES)
_http = AuthorizedHttp(creds, http=httplib2.Http(timeout=10))
_GCAL = build("calendar", "v3", http=_http, cache_discovery=False)
_SHEETS = build("sheets", "v4", http=_http, cache_discovery=False)

# --- DB helpers ---

async def db_exec(query: str, params: tuple = ()):
    d = os.path.dirname(DB_PATH)
    if d: os.makedirs(d, exist_ok=True)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(query, params)
        await db.commit()

async def db_fetchone(query: str, params: tuple = ()):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(query, params)
        row = await cur.fetchone()
        await cur.close()
        return row

async def db_fetchall(query: str, params: tuple = ()):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(query, params)
        rows = await cur.fetchall()
        await cur.close()
        return rows

async def ensure_db():
    await db_exec("""
        CREATE TABLE IF NOT EXISTS events_map(
          discord_message_id INTEGER PRIMARY KEY,
          event_id TEXT NOT NULL,
          channel_id INTEGER NOT NULL,
          thread_id INTEGER
        )""")
    await db_exec("""
        CREATE TABLE IF NOT EXISTS rsvps(
          event_id TEXT NOT NULL,
          user_id INTEGER NOT NULL,
          status TEXT NOT NULL,
          PRIMARY KEY (event_id, user_id)
        )""")
    await db_exec("""
        CREATE TABLE IF NOT EXISTS event_tags(
          event_id TEXT NOT NULL,
          tag TEXT NOT NULL,
          PRIMARY KEY (event_id, tag)
        )""")


# optional gspread helpers if you installed gspread
try:
    import gspread
    def get_sheets_client():
        return gspread.authorize(creds)
    def open_ws(sheet_id: str, tab: str | None = None):
        sh = get_sheets_client().open_by_key(sheet_id)
        return sh.worksheet(tab) if tab else sh.sheet1
except Exception:
    pass

def display_dt(dt_iso: str | None) -> str:
    if not dt_iso:
        return ""
    dt = du_parser.isoparse(dt_iso)
    local = dt.astimezone(TZ)
    return local.strftime("%a, %b %d at %I:%M %p %Z")

# --- Sheets helpers (use your tab names) ---
def _a1(tab: str, rng: str) -> str:
    safe = (tab or "").replace("'", "''")
    return f"'{safe}'!{rng}" if tab else rng

def _sheets_get(sheet_id: str, a1: str, retries: int = 3) -> list[list[str]]:
    for attempt in range(1, retries + 1):
        try:
            resp = _SHEETS.spreadsheets().values().get(
                spreadsheetId=sheet_id, range=a1
            ).execute(num_retries=2)
            return resp.get("values", []) or []
        except Exception:
            if attempt == retries:
                raise
            time.sleep((2 ** (attempt - 1)) + random.random())

async def sheet_values(sheet_id: str, tab: str, rng: str) -> list[list[str]]:
    a1 = _a1(tab, rng)
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, _sheets_get, sheet_id, a1)

def _rows_to_dicts(values: list[list[str]]) -> list[dict]:
    if not values: return []
    hdr = [h.strip() for h in values[0]]
    out = []
    for row in values[1:]:
        row = row + [""] * (len(hdr) - len(row))
        out.append(dict(zip(hdr, row)))
    return out

# Read once, return {role_type: [role names]}
async def list_roles_by_type(allowed: Optional[set[str]] = None) -> Dict[str, List[Tuple[str, str]]]:
    """
    Returns {role_type: [(name, desc), ...]} from the Roles sheet.
    Expected headers: Role name | Role Type | Description (case-insensitive)
    If Description is missing, use empty string.
    """
    vals = await sheet_values(ROLES_SHEET, ROLES_TAB, "A:C")
    if not vals:
        return {}

    hdr = [h.strip().lower() for h in vals[0]]
    try:
        i_role = hdr.index("role name")
    except ValueError:
        i_role = 0
    try:
        i_type = hdr.index("role type")
    except ValueError:
        i_type = 1
    i_desc = hdr.index("description") if "description" in hdr else None

    out: Dict[str, List[Tuple[str, str]]] = {}
    for r in vals[1:]:
        if len(r) <= max(i_role, i_type):
            continue
        t = str(r[i_type]).strip()
        if allowed and t not in allowed:
            continue
        name = str(r[i_role]).strip()
        if not name:
            continue
        desc = str(r[i_desc]).strip() if i_desc is not None and i_desc < len(r) else ""
        out.setdefault(t, []).append((name, desc))
    return out

async def get_tags(tag_type: str) -> List[Tuple[str, str]]:
    """Return [(name, desc)] for a single Role Type."""
    data = await list_roles_by_type(allowed={tag_type})
    return data.get(tag_type, [])


# Convenience: get only one or many types
async def list_roles_for(types: list[str] | set[str]) -> list[str]:
    m = await list_roles_by_type(set(t.strip().lower() for t in types))
    # flatten, keep order within each type
    return [n for _, names in m.items() for n in names]

# Backwards-compat: interests only
async def list_interest_roles() -> list[str]:
    return await list_roles_for({"interest"})

# Discover available types from sheet
async def list_role_types() -> list[str]:
    m = await list_roles_by_type(None)
    return sorted(m.keys())


# members tab (same spreadsheet)
async def list_members() -> list[dict]:
    if not ROLES_SHEET or not MEMBERS_TAB:
        return []
    vals = await sheet_values(ROLES_SHEET, MEMBERS_TAB, "A:Z")
    return _rows_to_dicts(vals)

def norm_tag(t: str) -> str:
    return t.strip()
# stricter:
# def norm_tag(t: str) -> str:
#     return " ".join(t.split()).strip().lower()

# --- Excel helpers (local .xlsx) ---

def _ensure_parent(path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    return p

def _to_dataframe(
    rows: Iterable[Mapping[str, Any]] | Iterable[Iterable[Any]],
    headers: list[str] | None = None,
) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=headers or [])
    first = next(iter(rows))
    # If dict-like, build from records
    if isinstance(first, Mapping):
        df = pd.DataFrame(list(rows))
    else:
        df = pd.DataFrame(list(rows), columns=headers)
    return df

def _autosize_excel_columns(ws) -> None:
    from openpyxl.utils import get_column_letter
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells:
            v = c.value
            try:
                l = len(str(v)) if v is not None else 0
            except Exception:
                l = 0
            if l > max_len:
                max_len = l
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

def write_excel_table(
    path: str | Path,
    sheet_name: str,
    rows: Iterable[Iterable[Any]],
    headers: list[str] | None = None,
    mode: str = "replace",  # "replace" | "append"
    index: bool = False,
) -> str:
    """
    Write 2D rows to an .xlsx sheet.
    - replace: overwrite the sheet
    - append:   append below existing data (requires headers match)
    Returns the absolute path.
    """
    p = _ensure_parent(path)
    df = _to_dataframe(rows, headers)

    if mode not in {"replace", "append"}:
        raise ValueError("mode must be 'replace' or 'append'")

    if not p.exists() or mode == "replace":
        with pd.ExcelWriter(p, engine="openpyxl") as xw:
            df.to_excel(xw, sheet_name=sheet_name, index=index)
            ws = xw.book[sheet_name]
            _autosize_excel_columns(ws)
        return str(p.resolve())

    # append
    from openpyxl import load_workbook
    wb = load_workbook(p)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # write headers + data
        ws.append(list(df.columns))
        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))
        _autosize_excel_columns(ws)
        wb.save(p)
        return str(p.resolve())

    ws = wb[sheet_name]
    # Read existing header to align columns
    existing_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if existing_headers and list(df.columns) != existing_headers:
        raise ValueError("append failed: headers do not match existing sheet")
    if not existing_headers:
        ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    _autosize_excel_columns(ws)
    wb.save(p)
    return str(p.resolve())

def write_excel_dicts(
    path: str | Path,
    sheet_name: str,
    records: Iterable[Mapping[str, Any]],
    mode: str = "replace",  # "replace" | "append"
    index: bool = False,
) -> str:
    """
    Write list[dict] to an .xlsx sheet. Keys become headers.
    Returns the absolute path.
    """
    # Build a stable header order across records
    cols: list[str] = []
    rec_list = list(records)
    for r in rec_list:
        for k in r.keys():
            if k not in cols:
                cols.append(k)
    rows = ([r.get(c) for c in cols] for r in rec_list)
    return write_excel_table(path, sheet_name, rows, headers=cols, mode=mode, index=index)

def update_excel_row(
    path: str | Path,
    sheet_name: str,
    match_col: str,
    match_val,
    updates: dict,
    *,
    update_first_only: bool = False,
    insert_if_missing: bool = True,
    add_missing_columns: bool = True,
) -> tuple[int, int, str]:
    """
    Update rows where df[match_col] == match_val with values in `updates`.
    If no match and insert_if_missing=True, append a new row.
    Returns (rows_updated, rows_inserted, absolute_path).
    """
    p = _ensure_parent(path)
    rows_updated = 0
    rows_inserted = 0

    # Load existing sheet -> DataFrame (or empty)
    if p.exists():
        try:
            df = pd.read_excel(p, sheet_name=sheet_name, dtype=object)
        except ValueError:
            # workbook exists, sheet doesn't
            df = pd.DataFrame(columns=[match_col])
    else:
        df = pd.DataFrame(columns=[match_col])

    # Ensure match column exists
    if match_col not in df.columns:
        df[match_col] = pd.Series(dtype=object)

    # Ensure update columns exist
    if add_missing_columns:
        for c in updates.keys():
            if c not in df.columns:
                df[c] = pd.Series(dtype=object)

    # Find matches
    mask = (df[match_col] == match_val)
    if mask.any():
        idx = df[mask].index
        if update_first_only:
            idx = [idx[0]]
        for col, val in updates.items():
            if col not in df.columns and not add_missing_columns:
                continue
            if col not in df.columns and add_missing_columns:
                df[col] = pd.Series(dtype=object)
            df.loc[idx, col] = val
        rows_updated = len(idx)
    elif insert_if_missing:
        # Build new row with all columns present
        new_row = {c: None for c in df.columns}
        new_row[match_col] = match_val
        for col, val in updates.items():
            if col not in df.columns and add_missing_columns:
                df[col] = pd.Series(dtype=object)
            new_row[col] = val
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        rows_inserted = 1
    else:
        # No match and no insert: nothing to do
        return (0, 0, str(p.resolve()))

    # Write back just this sheet, preserve others
    with pd.ExcelWriter(p, engine="openpyxl", mode="a" if p.exists() else "w",
                        if_sheet_exists="replace") as xw:
        df.to_excel(xw, sheet_name=sheet_name, index=False)
        ws = xw.book[sheet_name]
        _autosize_excel_columns(ws)

    return (rows_updated, rows_inserted, str(p.resolve()))

# --- Google Calendar helpers ---

async def gcal_insert_event(summary: str, start_dt, end_dt, location: str|None, description: str|None):
    body = {
        "summary": summary,
        "location": location or None,
        "description": description or None,
        "start": {"dateTime": start_dt.isoformat(), "timeZone": TZ_NAME},
        "end":   {"dateTime": end_dt.isoformat(),   "timeZone": TZ_NAME},
    }
    return await asyncio.to_thread(
        lambda: _GCAL.events().insert(calendarId=CAL_ID, body=body).execute()
    )


async def gcal_list(time_min_iso: str, time_max_iso: str, max_items: int = 25):
    return await asyncio.to_thread(
        lambda: _GCAL.events()
        .list(
            calendarId=CAL_ID,
            timeMin=time_min_iso,
            timeMax=time_max_iso,
            singleEvents=True,
            orderBy="startTime",
            maxResults=max_items,
        )
        .execute()
    )





